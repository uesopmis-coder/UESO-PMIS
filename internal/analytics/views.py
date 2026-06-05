from django.shortcuts import render
from django.utils import timezone
from datetime import datetime, timedelta
from . import services 

# Imports for Excel Export
import io
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from openpyxl.utils import get_column_letter 

# ==============================================================================
# HELPER FUNCTION (MOVED TO TOP LEVEL)
# ==============================================================================

def _get_validated_dates(request):
    """
    Helper function to get and validate start_date and end_date from request.
    Returns aware datetimes.
    """
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    current_tz = timezone.get_current_timezone()
    
    # Default to last 30 days if no dates are provided
    if not start_date_str or not end_date_str:
        end_date = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        start_date = end_date - timedelta(days=29)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        
        context_dates = {
            'selected_start_date': start_date,
            'selected_end_date': end_date,
        }
        return start_date, end_date, context_dates

    # Parse dates if provided
    try:
        start_date_naive = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date_naive = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        # Make aware, setting start to beginning of day and end to end of day
        start_date = current_tz.localize(start_date_naive.replace(hour=0, minute=0, second=0, microsecond=0))
        end_date = current_tz.localize(end_date_naive.replace(hour=23, minute=59, second=59, microsecond=999999))
        
        context_dates = {
            'selected_start_date': start_date,
            'selected_end_date': end_date,
        }
        return start_date, end_date, context_dates
        
    except (ValueError, TypeError):
        end_date = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        start_date = end_date - timedelta(days=29)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        
        context_dates = {
            'selected_start_date': start_date,
            'selected_end_date': end_date,
        }
        return start_date, end_date, context_dates

# ==============================================================================
# VIEWS
# ==============================================================================

def analytics_view(request):
    
    start_date, end_date, context_dates = _get_validated_dates(request)
    
    user_college = None
    is_college_restricted = False
    if hasattr(request.user, 'role') and request.user.role in ['PROGRAM_HEAD', 'DEAN', 'COORDINATOR']:
        user_college = request.user.college
        is_college_restricted = True
    
    context = {
        'card_metrics': {
            'total_projects': services.get_total_projects_count(start_date, end_date, college=user_college),
            'total_events': services.get_total_events_count(start_date, end_date, college=user_college),
            'total_providers': services.get_total_providers_count(start_date, end_date, college=user_college),
            'total_trained': services.get_total_individuals_trained(start_date, end_date, college=user_college),
        },
        'project_trends': services.get_project_trends(start_date, end_date, college=user_college),
        'is_college_restricted': is_college_restricted,  
    }
    
    context.update(context_dates)

    return render(request, 'analytics.html', context)


def export_analytics_to_excel(request):
    """
    Gathers all analytics data for the given date range and exports it
    as a multi-sheet Excel file, compatible with the new multi-series budget chart.
    """
    start_date, end_date, _ = _get_validated_dates(request)
    
    user_college = None
    if hasattr(request.user, 'role') and request.user.role in ['PROGRAM_HEAD', 'DEAN', 'COORDINATOR']:
        user_college = request.user.college
    
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)

    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    ws_overview['A1'] = "Analytics Report"
    ws_overview['A1'].font = title_font
    ws_overview['A2'] = f"Date Range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    
    ws_overview['A4'] = "Key Metrics"
    ws_overview['A4'].font = header_font
    
    metrics = {
        "Total Projects Started": services.get_total_projects_count(start_date, end_date, college=user_college)['metric'],
        "Total Events": services.get_total_events_count(start_date, end_date, college=user_college)['metric'],
        "Total Providers (Faculty & Colleges)": services.get_total_providers_count(start_date, end_date, college=user_college)['metric'],
        "Total Individuals Trained": services.get_total_individuals_trained(start_date, end_date, college=user_college)['metric']
    }
    
    row = 5
    for key, value in metrics.items():
        ws_overview[f'A{row}'] = key
        ws_overview[f'B{row}'] = value
        row += 1
    ws_overview.column_dimensions['A'].width = 35
    ws_overview.column_dimensions['B'].width = 15

    ws_projects = wb.create_sheet(title="Projects Over Time")
    ws_projects['A1'] = "Active Projects Created Over Time"
    ws_projects['A1'].font = header_font
    ws_projects['A3'] = "Date/Time Unit"
    ws_projects['B3'] = "Projects Created"
    ws_projects['A3'].font = header_font
    ws_projects['B3'].font = header_font
    
    project_data = services.get_active_projects_over_time(start_date, end_date, college=user_college).get('data', [])
    for i, item in enumerate(project_data, start=4):
        ws_projects[f'A{i}'] = item['x']
        ws_projects[f'B{i}'] = item['y']
    ws_projects.column_dimensions['A'].width = 25
    ws_projects.column_dimensions['B'].width = 20

    ws_trained = wb.create_sheet(title="Individuals Trained")
    ws_trained['A1'] = "Individuals Trained Over Time"
    ws_trained['A1'].font = header_font
    ws_trained['A3'] = "Date/Time Unit"
    ws_trained['B3'] = "Individuals Trained"
    ws_trained['A3'].font = header_font
    ws_trained['B3'].font = header_font
    
    trained_data = services.get_trained_individuals_data(start_date, end_date, college=user_college).get('data', [])
    for i, item in enumerate(trained_data, start=4):
        ws_trained[f'A{i}'] = item['x']
        ws_trained[f'B{i}'] = item['y']
    ws_trained.column_dimensions['A'].width = 25
    ws_trained.column_dimensions['B'].width = 20

    ws_budget = wb.create_sheet(title="Budget Allocation")
    ws_budget['A1'] = "Budget Allocation by College"
    ws_budget['A1'].font = title_font
    ws_budget['A2'] = f"Data Based on Fiscal Year: {end_date.year}"
    
    budget_data = services.get_budget_allocation_data(start_date, end_date, college=user_college)
    labels = budget_data.get('labels', [])
    datasets = budget_data.get('datasets', [])
    
    ws_budget['A4'] = "College"
    ws_budget['A4'].font = header_font
    
    current_col = 2
    for dataset in datasets:
        header_title = dataset['label']
        cell = ws_budget.cell(row=4, column=current_col, value=f"{header_title} (₱)")
        cell.font = header_font
        current_col += 1
        
    for i, label in enumerate(labels, start=5):
        ws_budget[f'A{i}'] = label
        
        current_col = 2
        for dataset in datasets:
            data_list = dataset['data']
            value = data_list[i-5] 
            cell = ws_budget.cell(row=i, column=current_col, value=value)
            cell.number_format = '₱#,##0.00'
            current_col += 1

    ws_budget.column_dimensions['A'].width = 40
    for col_idx in range(2, current_col):
        ws_budget.column_dimensions[get_column_letter(col_idx)].width = 25

    ws_agenda = wb.create_sheet(title="Agenda Distribution")
    ws_agenda['A1'] = "Project Distribution by Agenda"
    ws_agenda['A1'].font = header_font
    ws_agenda['A3'] = "Agenda"
    ws_agenda['B3'] = "Project Count"
    ws_agenda['A3'].font = header_font
    ws_agenda['B3'].font = header_font
    
    agenda_data = services.get_agenda_distribution_data(start_date, end_date, college=user_college)
    for i, label in enumerate(agenda_data.get('labels', []), start=4):
        ws_agenda[f'A{i}'] = label
        ws_agenda[f'B{i}'] = agenda_data['counts'][i-4]
    ws_agenda.column_dimensions['A'].width = 40
    ws_agenda.column_dimensions['B'].width = 20

    ws_requests = wb.create_sheet(title="Request Status")
    ws_requests['A1'] = "Client Request Status"
    ws_requests['A1'].font = header_font
    
    request_data = services.get_request_status_distribution(start_date, end_date, college=user_college)
    total_count = request_data.get('total_count', 0)
    
    ws_requests['A3'] = "Status"
    ws_requests['B3'] = "Percentage"
    ws_requests['C3'] = "Count (Approx.)"
    ws_requests['A3'].font = header_font
    ws_requests['B3'].font = header_font
    ws_requests['C3'].font = header_font
    
    ws_requests['A4'] = "Approved"
    ws_requests['A5'] = "Ongoing"
    ws_requests['A6'] = "Rejected"
    
    ws_requests['B4'] = request_data.get('approved_pct', 0) / 100
    ws_requests['B5'] = request_data.get('ongoing_pct', 0) / 100
    ws_requests['B6'] = request_data.get('rejected_pct', 0) / 100
    ws_requests['B4'].number_format = '0.0%'
    ws_requests['B5'].number_format = '0.0%'
    ws_requests['B6'].number_format = '0.0%'
    
    ws_requests['C4'] = round(total_count * (request_data.get('approved_pct', 0) / 100))
    ws_requests['C5'] = round(total_count * (request_data.get('ongoing_pct', 0) / 100))
    ws_requests['C6'] = round(total_count * (request_data.get('rejected_pct', 0) / 100))
    
    ws_requests['A8'] = "Total Requests"
    ws_requests['C8'] = total_count
    ws_requests['A8'].font = header_font
    ws_requests['C8'].font = header_font

    ws_requests.column_dimensions['A'].width = 20
    ws_requests.column_dimensions['B'].width = 15
    ws_requests.column_dimensions['C'].width = 20
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"analytics_export_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response