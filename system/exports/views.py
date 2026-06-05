from django.conf import settings
from django.core.paginator import Paginator
from django.core.mail import EmailMultiAlternatives
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from django.contrib import messages

from .models import ExportRequest, can_export_direct, must_request_export
from system.users.decorators import role_required
from system.users.models import User
from system.utils.email_utils import async_send_export_approved, async_send_export_rejected
from shared.projects.models import Project

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO


@role_required(allowed_roles=["UESO", "VP", "DIRECTOR"], require_confirmed=True)
def exports_view(request):
    # Optimize with select_related
    requests = ExportRequest.objects.select_related('submitted_by', 'reviewed_by').all()

    # Filters
    sort_by = request.GET.get('sort_by', 'date_submitted')
    order = request.GET.get('order', 'desc')
    status = request.GET.get('status', '')
    export_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '').strip()


    # Apply filters
    if status:
        requests = requests.filter(status__iexact=status)
    if export_type:
        requests = requests.filter(type=export_type)
    if date_from:
        requests = requests.filter(date_submitted__date__gte=date_from)
    if date_to:
        requests = requests.filter(date_submitted__date__lte=date_to)
    if search:
        from django.db.models import Q
        requests = requests.filter(type__icontains=search)

    requests = requests.distinct()

    # Sorting
    sort_map = {
        'date_submitted': 'date_submitted',
        'status': 'status',
        'type': 'type',
    }
    sort_field = sort_map.get(sort_by, 'date_submitted')
    if sort_field:
        if order == 'desc':
            sort_field = '-' + sort_field
        requests = requests.order_by(sort_field)

    # Filter Options
    all_statuses = ExportRequest._meta.get_field('status').choices
    all_types = ExportRequest._meta.get_field('type').choices

    # Pagination
    paginator = Paginator(requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    page_range = paginator.get_elided_page_range(page_obj.number)

    return render(request, 'exports/exports.html', {
        'search': search,
        'sort_by': sort_by,
        'order': order,
        'all_statuses': all_statuses,
        'all_types': all_types,
        'status': status,
        'export_type': export_type,
        'date_from': date_from,
        'date_to': date_to,
        'page_obj': page_obj,
        'page_range': page_range,
        'querystring': request.GET.urlencode().replace('&page='+str(page_obj.number), '') if page_obj else '',
    })


########################################################################################################################


@role_required(allowed_roles=["UESO", "VP", "DIRECTOR"], require_confirmed=True)
def approve_export_request(request, request_id):
    """Approve an export request and send email asynchronously"""
    try:
        export_request = ExportRequest.objects.select_related('submitted_by', 'reviewed_by').get(id=request_id, status='PENDING')
    except ExportRequest.DoesNotExist:
        messages.error(request, 'Export request not found or already processed.')
        return redirect('exports')
    
    # Set approval details
    export_request.status = 'APPROVED'
    export_request.reviewed_by = request.user
    export_request.reviewed_at = timezone.now()
    export_request.save()

    # Send email asynchronously with download link (non-blocking, no 2-minute delay!)
    user = export_request.submitted_by
    export_type_display = export_request.get_type_display()
    
    if user.email:
        # Build full download URL
        from django.urls import reverse
        scheme = 'https' if request.is_secure() else 'http'
        domain = request.get_host()
        download_path = reverse('export_download', args=[export_request.id])
        download_url = f"{scheme}://{domain}{download_path}"
        
        async_send_export_approved(user.email, export_type_display, download_url)
    
    messages.success(request, f'{export_type_display} export request approved. Email sent to {user.get_full_name()}.')
    
    # Redirect back to exports page
    return redirect('exports')


@require_POST
@role_required(allowed_roles=["UESO", "VP", "DIRECTOR"], require_confirmed=True)
def reject_export_request(request, request_id):
    """Reject an export request and send notification email asynchronously"""
    try:
        export_request = ExportRequest.objects.select_related('submitted_by').get(id=request_id, status='PENDING')
    except ExportRequest.DoesNotExist:
        messages.error(request, 'Export request not found or already processed.')
        return redirect('exports')
    
    # Set rejection details
    export_request.status = 'REJECTED'
    export_request.reviewed_by = request.user
    export_request.reviewed_at = timezone.now()
    export_type_display = export_request.get_type_display()
    export_request.save()
    
    # Send rejection email asynchronously (non-blocking)
    user = export_request.submitted_by
    if user.email:
        async_send_export_rejected(user.email, export_type_display)
    
    messages.success(request, f'{export_type_display} export request rejected. Email sent to {user.get_full_name()}.')
    
    # Redirect back to exports page
    return redirect('exports')


# Robust download endpoint for filtered export files
def export_download(request, request_id):
    export_request = get_object_or_404(ExportRequest, id=request_id, status='APPROVED')
    # Only allow the submitter or reviewers to download
    user = request.user
    allowed_roles = ["UESO", "VP", "DIRECTOR"]
    if not (user == export_request.submitted_by or (hasattr(user, 'role') and user.role in allowed_roles)):
        return JsonResponse({'error': 'You do not have permission to download this export.'}, status=403)

    from urllib.parse import parse_qs
    file_buffer = BytesIO()
    filename = None
    qs = parse_qs(export_request.querystring)
    
    # Get the original submitter to check their role-based restrictions
    submitter = export_request.submitted_by

    if export_request.type == 'MANAGE_USER':
        users = User.objects.all()
        from django.db.models import Q
        search = qs.get('search', [''])[0].strip()
        if search:
            users = users.filter(
                Q(given_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(middle_initial__icontains=search) |
                Q(suffix__icontains=search) |
                Q(email__icontains=search)
            )
        sort_by = qs.get('sort_by', ['date'])[0]
        order = qs.get('order', ['desc'])[0]
        role = qs.get('role', [''])[0]
        verified = qs.get('verified', [''])[0]
        date = qs.get('date', [''])[0]
        college = qs.get('college', [''])[0]
        campus = qs.get('campus', [''])[0]
        
        # Auto-filter by submitter's college if they have restricted role
        if submitter.role in ['PROGRAM_HEAD', 'DEAN', 'COORDINATOR'] and submitter.college:
            college = str(submitter.college.id)
        
        if role:
            users = users.filter(role=role)
        if verified == 'true':
            users = users.filter(is_confirmed=True)
        elif verified == 'false':
            users = users.filter(is_confirmed=False)
        if date:
            users = users.filter(date_joined__date=date)
        if college:
            users = users.filter(college_id=college)
        if campus:
            users = users.filter(college__campus_id=campus)
        if sort_by == 'name':
            sort_field = ['last_name', 'given_name', 'middle_initial', 'suffix']
        else:
            sort_map = {
                'email': 'email',
                'date': 'date_joined',
                'role': 'role',
            }
            sort_field = [sort_map.get(sort_by, 'last_name')]
        if order == 'desc':
            sort_field = ['-' + f for f in sort_field]
        users = users.order_by(*sort_field)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Manage Users"
        headers = [
            'Last Name', 'Given Name', 'Middle Initial', 'Suffix', 'Email', 'Role', 'Verified', 'Date Joined', 'College', 'Campus'
        ]
        ws.append(headers)
        for u in users:
            ws.append([
                u.last_name,
                u.given_name,
                u.middle_initial,
                u.suffix,
                u.email,
                u.get_role_display() if hasattr(u, 'get_role_display') else getattr(u, 'role', ''),
                u.is_confirmed,
                u.date_joined.strftime('%Y-%m-%d %H:%M'),
                str(getattr(u, 'college', '')),
                u.get_campus_display() if hasattr(u, 'get_campus_display') else getattr(u, 'campus', ''),
            ])
        wb.save(file_buffer)
        filename = 'manage_users_export.xlsx'
    elif export_request.type == 'PROJECT':
        projects = Project.objects.all()
        from django.db.models import Q
        search = qs.get('search', [''])[0].strip()
        sort_by = qs.get('sort_by', ['last_updated'])[0]
        order = qs.get('order', ['desc'])[0]
        college = qs.get('college', [''])[0]
        campus = qs.get('campus', [''])[0]
        agenda = qs.get('agenda', [''])[0]
        status = qs.get('status', [''])[0]
        year = qs.get('year', [''])[0]
        quarter = qs.get('quarter', [''])[0]
        date = qs.get('date', [''])[0]
        
        # Auto-filter by submitter's college if they have restricted role
        if submitter.role in ['PROGRAM_HEAD', 'DEAN', 'COORDINATOR'] and submitter.college:
            college = str(submitter.college.id)
        
        if college:
            projects = projects.filter(project_leader__college__id=college)
        if campus:
            projects = projects.filter(project_leader__college__campus_id=campus)
        if agenda:
            projects = projects.filter(agenda__id=agenda)
        if status:
            projects = projects.filter(status=status)
        if year:
            projects = projects.filter(start_date__year=year)
        if quarter:
            qmap = {'1': (1,3), '2': (4,6), '3': (7,9), '4': (10,12)}
            if quarter in qmap:
                start, end = qmap[quarter]
                projects = projects.filter(start_date__month__gte=start, start_date__month__lte=end)
        if date:
            projects = projects.filter(start_date=date)
        if search:
            projects = projects.filter(title__icontains=search)
        sort_map = {
            'title': 'title',
            'last_updated': 'updated_at',
            'start_date': 'start_date',
            'progress': '',
        }
        sort_field = sort_map.get(sort_by, 'title')
        if sort_field:
            if order == 'desc':
                sort_field = '-' + sort_field
            projects = projects.order_by(sort_field)
        elif sort_by == 'progress':
            projects = sorted(projects, key=lambda p: (p.progress[0] / p.progress[1]) if p.progress[1] else 0, reverse=(order=='desc'))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        headers = [
            'Title', 'Leader', 'College/Unit', 'Last Updated', 'Start Date', 'Progress', 'Status'
        ]
        ws.append(headers)
        for p in projects:
            ws.append([
                p.title,
                p.project_leader.get_full_name() if p.project_leader else '',
                p.project_leader.college.name if p.project_leader and p.project_leader.college else '',
                p.updated_at.strftime('%Y-%m-%d') if p.updated_at else '',
                p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                getattr(p, 'progress_display', ''),
                p.get_status_display() if hasattr(p, 'get_status_display') else p.status,
            ])
        wb.save(file_buffer)
        filename = 'projects_export.xlsx'

    # BUDGET
    # GOAL

    if filename:
        file_buffer.seek(0)
        return FileResponse(file_buffer, as_attachment=True, filename=filename, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return JsonResponse({'error': 'Export type not supported.'}, status=400)


########################################################################################################################


@require_GET
@role_required(allowed_roles=["UESO", "VP", "DIRECTOR"], require_confirmed=True)
def export_manage_user(request):
    user = request.user
    UserModel = User
    # Optimize with select_related for export
    users = UserModel.objects.select_related('college', 'college__campus').all()
    query_params = {}

    # --- Filters (match manage_user view) ---
    from django.db.models import Q
    search = request.GET.get('search', '').strip()
    if search:
        users = users.filter(
            Q(given_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(middle_initial__icontains=search) |
            Q(suffix__icontains=search) |
            Q(email__icontains=search)
        )
        query_params['search'] = search

    sort_by = request.GET.get('sort_by', 'date')
    order = request.GET.get('order', 'desc')
    role = request.GET.get('role', '')
    verified = request.GET.get('verified', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    college = request.GET.get('college', '')
    campus = request.GET.get('campus', '')

    # Auto-filter by college for PROGRAM_HEAD, DEAN, COORDINATOR (override any college filter)
    if user.role in ['PROGRAM_HEAD', 'DEAN', 'COORDINATOR'] and user.college:
        # Force college to be user's college - they can only see their own college data
        college = str(user.college.id)

    # Apply filters
    if sort_by:
        query_params['sort_by'] = sort_by
    if order:
        query_params['order'] = order
    if role:
        users = users.filter(role=role)
        query_params['role'] = role
    if verified == 'true':
        users = users.filter(is_confirmed=True)
        query_params['verified'] = 'true'
    elif verified == 'false':
        users = users.filter(is_confirmed=False)
        query_params['verified'] = 'false'
    if date_from:
        users = users.filter(date_joined__date__gte=date_from)
        query_params['date_from'] = date_from
    if date_to:
        users = users.filter(date_joined__date__lte=date_to)
        query_params['date_to'] = date_to
    if college:
        users = users.filter(college_id=college)
        query_params['college'] = college
    if campus:
        users = users.filter(college__campus_id=campus)
        query_params['campus'] = campus

    # Sorting
    if sort_by == 'name':
        sort_field = ['last_name', 'given_name', 'middle_initial', 'suffix']
    else:
        sort_map = {
            'email': 'email',
            'date': 'date_joined',
            'role': 'role',
        }
        sort_field = [sort_map.get(sort_by, 'last_name')]
    if order == 'desc':
        sort_field = ['-' + f for f in sort_field]
    users = users.order_by(*sort_field)

    # Generate XLSX (direct export)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manage Users"
    headers = [
        'Last Name', 'Given Name', 'Middle Initial', 'Suffix', 'Email', 'Role', 'Verified', 'Date Joined', 'College', 'Campus'
    ]
    ws.append(headers)
    for u in users:
        ws.append([
            u.last_name,
            u.given_name,
            u.middle_initial,
            u.suffix,
            u.email,
            u.get_role_display() if hasattr(u, 'get_role_display') else getattr(u, 'role', ''),
            u.is_confirmed,
            u.date_joined.strftime('%Y-%m-%d %H:%M'),
            str(getattr(u, 'college', '')),
            u.get_campus_display() if hasattr(u, 'get_campus_display') else getattr(u, 'campus', ''),
        ])
    # Auto-fit column widths
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 12)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="manage_users_export.xlsx"'
    return response


@require_GET
@role_required(allowed_roles=["UESO", "VP", "DIRECTOR", "DEAN", "PROGRAM_HEAD", "COORDINATOR"], require_confirmed=True)
def export_project(request):
    user = request.user
    from django.db.models import Q
    # Optimize with select_related and prefetch_related for export
    projects = Project.objects.select_related(
        'project_leader',
        'project_leader__college',
        'project_leader__college__campus',
        'agenda'
    ).prefetch_related('providers', 'sdgs', 'events').all()
    # Filters (match admin_project view)
    search = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort_by', 'last_updated')
    order = request.GET.get('order', 'desc')
    college = request.GET.get('college', '')
    campus = request.GET.get('campus', '')
    agenda = request.GET.get('agenda', '')
    status = request.GET.get('status', '')
    year = request.GET.get('year', '')
    quarter = request.GET.get('quarter', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Auto-filter by college for PROGRAM_HEAD, DEAN, COORDINATOR (override any college filter)
    if user.role in ['PROGRAM_HEAD', 'DEAN', 'COORDINATOR'] and user.college:
        # Force college to be user's college - they can only see their own college data
        college = user.college.id

    # Apply all filters
    if college:
        projects = projects.filter(project_leader__college__id=college)
    if campus:
        projects = projects.filter(project_leader__college__campus_id=campus)
    if agenda:
        projects = projects.filter(agenda__id=agenda)
    if status:
        projects = projects.filter(status=status)
    if year:
        projects = projects.filter(start_date__year=year)
    if quarter:
        # Q1: Jan-Mar, Q2: Apr-Jun, Q3: Jul-Sep, Q4: Oct-Dec
        qmap = {'1': (1,3), '2': (4,6), '3': (7,9), '4': (10,12)}
        if quarter in qmap:
            start, end = qmap[quarter]
            projects = projects.filter(start_date__month__gte=start, start_date__month__lte=end)
    if date_from:
        projects = projects.filter(start_date__gte=date_from)
    if date_to:
        projects = projects.filter(start_date__lte=date_to)
    if search:
        projects = projects.filter(title__icontains=search)

    # Sorting
    sort_map = {
        'title': 'title',
        'last_updated': 'updated_at',
        'start_date': 'start_date',
        'progress': '', # Placeholder, not supported in DB sort
    }
    sort_field = sort_map.get(sort_by, 'title')
    if sort_field:
        if order == 'desc':
            sort_field = '-' + sort_field
        projects = projects.order_by(sort_field)
    # If progress sort, sort in Python (not supported in DB)
    elif sort_by == 'progress':
        projects = sorted(projects, key=lambda p: (p.progress[0] / p.progress[1]) if p.progress[1] else 0, reverse=(order=='desc'))

    if can_export_direct(user):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        headers = [
            'Title', 'Leader', 'College/Unit', 'Last Updated', 'Start Date', 'Progress', 'Status'
        ]
        ws.append(headers)
        for p in projects:
            ws.append([
                p.title,
                p.project_leader.get_full_name() if p.project_leader else '',
                p.project_leader.college.name if p.project_leader and p.project_leader.college else '',
                p.updated_at.strftime('%Y-%m-%d') if p.updated_at else '',
                p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                getattr(p, 'progress_display', ''),
                p.get_status_display() if hasattr(p, 'get_status_display') else p.status,
            ])
        # Auto-fit column widths
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 12)
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="projects_export.xlsx"'
        return response
    elif must_request_export(user):
        ExportRequest.objects.create(
            type='PROJECT',
            date_submitted=timezone.now(),
            submitted_by=user,
            status='PENDING',
            querystring=request.META.get('QUERY_STRING', '')
        )
        return JsonResponse({'message': 'Your export request has been submitted for approval.'}, status=202)
    else:
        return JsonResponse({'error': 'You do not have permission to export.'}, status=403)



from system.logs.models import LogEntry
from system.users.models import User

@require_GET
@role_required(allowed_roles=["VP", "DIRECTOR"], require_confirmed=True)
def export_log(request):
    # Get filter parameters (match logs_view)
    sort_by = request.GET.get('sort_by', 'timestamp')
    order = request.GET.get('order', 'desc')
    user_role = request.GET.get('user_role', '')
    action = request.GET.get('action', '')
    model = request.GET.get('model', '')
    search = request.GET.get('search', '').strip()

    logs = LogEntry.objects.select_related('user')
    if user_role:
        logs = logs.filter(user__role=user_role)
    if action:
        logs = logs.filter(action=action)
    if model:
        logs = logs.filter(model=model)
    if search:
        from django.db.models import Q
        logs = logs.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(object_repr__icontains=search)
        )

    sort_map = {
        'timestamp': 'timestamp',
        'user': 'user__first_name',
        'action': 'action',
        'model': 'model',
        'object_id': 'object_id',
        'object_repr': 'object_repr',
    }
    sort_field = sort_map.get(sort_by, 'timestamp')
    if order == 'desc':
        sort_field = '-' + sort_field
    logs = logs.order_by(sort_field)

    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs"
    headers = [
        'User', 'User Email', 'Action', 'Model', 'Object ID', 'Object Repr', 'Timestamp', 'Details', 'URL'
    ]
    ws.append(headers)
    for log in logs:
        ws.append([
            log.user.get_full_name() if log.user else '-',
            log.user.email if log.user and log.user.email else '-',
            log.get_action_display(),
            log.model,
            log.object_id,
            log.object_repr,
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
            log.details,
            log.url,
        ])
    # Auto-fit column widths
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 12)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="logs_export.xlsx"'
    return response

@require_GET
def export_goals(request):
    # Later
    return 0


@require_GET
def export_budget(request):
    # Later
    return 0